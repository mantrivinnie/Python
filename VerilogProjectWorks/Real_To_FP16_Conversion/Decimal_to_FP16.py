import struct
import numpy as np
import csv
import os

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def decimal_to_hex(value):
    """
    Convert a single decimal value to IEEE 754 16-bit hex.
    
    Args:
        value (float): Decimal number to convert
        
    Returns:
        str: Hexadecimal FP16 representation (e.g., '0x3C00')
    """
    fp16_value = np.float16(value)
    fp16_bytes = struct.pack('>e', fp16_value)
    fp16_int = struct.unpack('>H', fp16_bytes)[0]
    return '0x' + format(fp16_int, '04X')


def decimals_to_hex_array(decimal_array):
    """
    Convert an array of decimal values to IEEE 754 16-bit hex array.
    
    Args:
        decimal_array (list): List of decimal numbers
        
    Returns:
        list: List of hexadecimal FP16 values
    """
    return [decimal_to_hex(val) for val in decimal_array]


def hex_to_decimal(hex_value):
    """
    Convert IEEE 754 16-bit hex back to decimal.
    
    Args:
        hex_value (str): Hexadecimal FP16 value (e.g., '0x3C00')
        
    Returns:
        float: Decimal representation
    """
    bits_int = int(hex_value, 16)
    bytes_data = struct.pack('>H', bits_int)
    return struct.unpack('>e', bytes_data)[0]


def convert_decimal_to_fp16(decimal_input):
    """
    Convert decimal value(s) to FP16 hexadecimal.
    
    Args:
        decimal_input (float or list): Single decimal or array of decimals
        
    Returns:
        str or list: Single hex value or list of hex values
    """
    if isinstance(decimal_input, (list, tuple, np.ndarray)):
        return decimals_to_hex_array(decimal_input)
    else:
        return decimal_to_hex(decimal_input)


def convert_fp16_to_decimal(hex_input):
    """
    Convert FP16 hexadecimal back to decimal.
    
    Args:
        hex_input (str or list): Single hex value or list of hex values
        
    Returns:
        float or list: Single decimal or list of decimals
    """
    if isinstance(hex_input, (list, tuple)):
        return [hex_to_decimal(h) for h in hex_input]
    else:
        return hex_to_decimal(hex_input)


def read_csv_input(csv_file):
    """
    Read decimal values from a CSV file.
    
    Args:
        csv_file (str): Path to the CSV file
        
    Returns:
        list: List of decimal values
    """
    decimal_values = []
    with open(csv_file, 'r') as file:
        csv_reader = csv.reader(file)
        for row in csv_reader:
            for value in row:
                try:
                    decimal_values.append(float(value.strip()))
                except ValueError:
                    print(f"Warning: Skipping non-numeric value '{value}'")
    return decimal_values


def read_excel_input(excel_file):
    """
    Read decimal values from an Excel file (.xlsx).
    
    Args:
        excel_file (str): Path to the Excel file
        
    Returns:
        list: List of decimal values
        
    Raises:
        ImportError: If openpyxl is not installed
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel support. Install with: pip install openpyxl")
    
    from openpyxl import load_workbook
    decimal_values = []
    
    wb = load_workbook(excel_file)
    ws = wb.active
    
    for row in ws.iter_rows(values_only=True):
        for cell_value in row:
            if cell_value is not None:
                try:
                    decimal_values.append(float(cell_value))
                except (ValueError, TypeError):
                    print(f"Warning: Skipping non-numeric value '{cell_value}'")
    
    return decimal_values


def write_csv_output(hex_values, output_file, decimal_values=None):
    """
    Write hex values to a CSV file.
    
    Args:
        hex_values (list): List of hexadecimal FP16 values
        output_file (str): Path to the output CSV file
        decimal_values (list, optional): Corresponding decimal values for reference
    """
    with open(output_file, 'w', newline='') as file:
        csv_writer = csv.writer(file)
        
        # Write header
        if decimal_values:
            csv_writer.writerow(['Decimal', 'FP16 Hex'])
            for dec, hex_val in zip(decimal_values, hex_values):
                csv_writer.writerow([dec, hex_val])
        else:
            csv_writer.writerow(['FP16 Hex'])
            for hex_val in hex_values:
                csv_writer.writerow([hex_val])
    
    print(f"\nOutput saved to: {output_file}")


if __name__ == "__main__":
    print("=" * 80)
    print("IEEE 754 16-bit Half-Precision Float Converter")
    print("Excel/CSV Input to FP16 Hex CSV Output Converter")
    print("=" * 80)
    
    while True:
        print("\n--- Options ---")
        print("1. Convert from Excel (.xlsx) file")
        print("2. Convert from CSV file")
        print("3. Manual input (comma-separated)")
        print("4. Exit")
        
        choice = input("\nSelect option (1-4): ").strip()
        
        if choice == '1':
            # Excel file input
            if not EXCEL_AVAILABLE:
                print("\nError: openpyxl is not installed.")
                print("Install it with: pip install openpyxl")
                continue
            
            excel_input_file = input("\nEnter input Excel file path (.xlsx): ").strip()
            
            if not os.path.exists(excel_input_file):
                print(f"Error: File '{excel_input_file}' not found.")
                continue
            
            try:
                # Read decimal values from Excel
                decimal_array = read_excel_input(excel_input_file)
                
                if not decimal_array:
                    print("Error: No numeric values found in Excel file.")
                    continue
                
                # Convert to FP16 hex
                hex_array = convert_decimal_to_fp16(decimal_array)
                
                # Generate output filename
                base_name = os.path.splitext(excel_input_file)[0]
                output_file = f"{base_name}_FP16_HEX.csv"
                
                # Write to CSV output file
                write_csv_output(hex_array, output_file, decimal_array)
                
                # Display results
                print("\n" + "=" * 80)
                print("CONVERSION RESULTS")
                print("=" * 80)
                print(f"Input Decimal Values:   {len(decimal_array)} values read")
                print(f"Output FP16 Hex Values: {len(hex_array)} values written")
                
                # Show first few mappings
                print("\n--- Sample Conversion ---")
                for i in range(min(5, len(decimal_array))):
                    print(f"  [{i}] {decimal_array[i]:12.6f} -> {hex_array[i]}")
                
                if len(decimal_array) > 5:
                    print(f"  ... ({len(decimal_array) - 5} more values)")
                
                print("\n" + "=" * 80)
                
            except Exception as e:
                print(f"\nError processing Excel file: {e}")
        
        elif choice == '2':
            # CSV file input
            csv_input_file = input("\nEnter input CSV file path: ").strip()
            
            if not os.path.exists(csv_input_file):
                print(f"Error: File '{csv_input_file}' not found.")
                continue
            
            try:
                # Read decimal values from CSV
                decimal_array = read_csv_input(csv_input_file)
                
                if not decimal_array:
                    print("Error: No numeric values found in CSV file.")
                    continue
                
                # Convert to FP16 hex
                hex_array = convert_decimal_to_fp16(decimal_array)
                
                # Generate output filename
                base_name = os.path.splitext(csv_input_file)[0]
                output_file = f"{base_name}_FP16_HEX.csv"
                
                # Write to CSV output file
                write_csv_output(hex_array, output_file, decimal_array)
                
                # Display results
                print("\n" + "=" * 80)
                print("CONVERSION RESULTS")
                print("=" * 80)
                print(f"Input Decimal Values:   {len(decimal_array)} values read")
                print(f"Output FP16 Hex Values: {len(hex_array)} values written")
                
                # Show first few mappings
                print("\n--- Sample Conversion ---")
                for i in range(min(5, len(decimal_array))):
                    print(f"  [{i}] {decimal_array[i]:12.6f} -> {hex_array[i]}")
                
                if len(decimal_array) > 5:
                    print(f"  ... ({len(decimal_array) - 5} more values)")
                
                print("\n" + "=" * 80)
                
            except Exception as e:
                print(f"\nError processing CSV file: {e}")
        
        elif choice == '3':
            # Manual input
            print("\n--- Input Decimal Array ---")
            print("Enter decimal values separated by commas (e.g., 0.0, 1.5, -2.5, 3.14)")
            
            user_input = input("\nEnter decimal array: ").strip()
            
            try:
                # Parse user input
                decimal_strings = user_input.split(',')
                decimal_array = [float(val.strip()) for val in decimal_strings]
                
                # Convert to FP16 hex
                hex_array = convert_decimal_to_fp16(decimal_array)
                
                # Display results
                print("\n" + "=" * 80)
                print("CONVERSION RESULTS")
                print("=" * 80)
                print(f"\nInput Decimal Array:    {decimal_array}")
                print(f"Output FP16 Hex Array:  {hex_array}")
                
                # Show detailed mapping
                print("\n--- Detailed Conversion ---")
                for i, (dec, hex_val) in enumerate(zip(decimal_array, hex_array)):
                    print(f"  [{i}] {dec:12.6f} -> {hex_val}")
                
                # Ask to save to CSV
                save_option = input("\nSave to CSV file? (yes/no): ").strip().lower()
                if save_option in ['yes', 'y']:
                    output_file = input("Enter output CSV filename (default: output_FP16_HEX.csv): ").strip()
                    if not output_file:
                        output_file = "output_FP16_HEX.csv"
                    write_csv_output(hex_array, output_file, decimal_array)
                
                print("\n" + "=" * 80)
                
            except ValueError as e:
                print(f"\nError: Invalid input. Please enter numbers separated by commas.")
                print(f"Details: {e}")
            except Exception as e:
                print(f"\nUnexpected error: {e}")
        
        elif choice == '4':
            print("Exiting...")
            break
        else:
            print("Invalid option. Please select 1, 2, 3, or 4.")